# test_data package/data_login_page

import openpyxl


class LoginData:
    # create var containing list with data where each set of data is a dictionary {"key":"value"}
    data_login = [{"login": "standard_user", "password": "fake_sauce"},
                  {"login": "problem_user", "password": "secret_sauce"},
                  {"login": "standard_user", "password": "secret_sauce"}]

    # declare func as @staticmethod to access it from outside of class just with ClassName.funk_name
    # no (self) because method is statick
    @staticmethod
    def get_one_set_data(test_case_name):
        # create dictionary to extract data from excel
        login_data_dict = {}

        # create war and assign to it path to the Excel document with openpyxl.load_workbook() method
        book = openpyxl.load_workbook("C:\\Disk D\\Draft\\QA Tester\\Automation projects\\PyTestData1.xlsx")

        # work with active sheet in the Excel document
        sheet = book.active
        print(sheet)  # demo for git

        # extract all values without titles
        # in range(2, ...) - range starts from 2 because we don't need the title of the row or column but values only
        for i in range(2, sheet.max_row + 1):  # to get rows (runs from up to down)
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # to get columns (left to right)
                    login_data_dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column=j).value)
        return[login_data_dict]
        # despite we created dictionary to have keys, our test excepts list[], so we wrap dictionary into list
